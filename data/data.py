# -*- coding: utf-8 -*-
# Time : 2023/12/27 16:19
# Author : Lucy
# File : data.py
# Desc :
class ReadWrite:
    def __init__(self):
        self.txtpath=''
        self.excelpath=r'E:\2_TestingSoftware\0_Pythontest\data\test.xlsx'
        self.yamlpath=''

    def txtread(self):
        list1 = []
        with open(self.txtpath,'r',encoding='utf-8') as f:
            values=f.readlines()
            f.close()
        for data in values:
            data_v=data.strip('\n')
            list1.append(data_v)
        return list1

    def txtwrite(self,username,password):
        f=open(self.txtpath,'a',encoding='utf-8')
        values=f'{username},{password}\n'
        f.writelines(values)
        f.close()

    def excelread(self, sheetname):
        import openpyxl
        wb = openpyxl.load_workbook(self.excelpath)
        table = wb[sheetname]
        rows = table.max_row
        cols = table.max_column
        list2 = []
        for row in range(2, rows + 1):
            list1 = []
            for col in range(1, cols + 1):
                values = table.cell(row, col).value
                list1.append(values)
            list2.append(list1)
        return list2

    def excelwrite(self, *args, sheetname):
        import openpyxl
        wb = openpyxl.load_workbook(self.excelpath)
        table = wb[sheetname]
        rows = table.max_row
        cols = len(args)
        for col in range(cols):
            table.cell(rows + 1, col + 1).value = args[col]
        wb.save(self.excelpath)

    def mysqlread(self):
        import pymysql
        db = pymysql.connect(host='localhost', port=3306, user='root', password='root', database='T31', charset='utf8')
        cur = db.cursor()
        sql = 'select * from users where username="tester001"'
        cur.execute(sql)
        content = cur.fetchall()
        return content

    def mysqlwrite(self, username, password):
        import pymysql
        db = pymysql.connect(host='localhost', port=3306, user='root', password='root', database='T31', charset='utf8')
        cur = db.cursor()
        sql = f'insert into users values("{username}","{password}")'
        cur.execute(sql)
        db.commit()

    def yamlread(self):
        import yaml
        f = open(self.yamlpath, 'r', encoding='utf-8')
        content = f.read()
        data = yaml.safe_load(content)
        return data

    def yamlwrite(self, username, password):
        import yaml
        f = open(self.yamlpath, 'a', encoding='utf-8')
        data = {'username': username, 'password': password}
        yaml.safe_dump(data, f)
        f.close()